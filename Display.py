import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
import time
import pandas as pd
from datetime import datetime
import tkinter.messagebox as messagebox
import requests
import pyodbc
import threading
from openpyxl import load_workbook 
import os
def create_app():
    api_url= 'xxxx'
    factory_name= '明治'
    value_bolean = 1
    root = tk.Tk()
    button_bg_color = "#d3d3d3" 
    ##d3d3d3 // mau xam
    #4da6ff //xanh duong
    #FF5722 mau cam

    root.title("生産管理版")

    # Tạo nhãn cho tiêu đề
    title_label = tk.Label(root, text="生　産　管　理　版", font=("Helvetica", 20, "bold"))
    title_label.pack(pady=10)

    root.tk.call("source", "azure.tcl")
    root.tk.call("set_theme", "dark")

    # Tạo nhãn cho thông tin chung
    info_frame = tk.Frame(root)
    info_frame.pack(pady=10) 

    tk.Label(info_frame, text="車種").grid(row=0, column=0, padx=10)
    car_type_combobox = ttk.Combobox(info_frame)
    car_type_combobox.grid(row=0, column=1, padx=10)    



    tk.Label(info_frame, text="工場").grid(row=0, column=2, padx=10)
    # Tạo combobox cho 車種
    factory_types = ["平子", "明治川","明海"]
    factory_type_combobox = ttk.Combobox(info_frame, values=factory_types)
    factory_type_combobox.grid(row=0, column=3, padx=10)
    factory_type_combobox.current(1) 

    # Tạo DateEntry cho 日付
    single_date_label = tk.Label(info_frame, text="日付")
    single_date_entry = DateEntry(info_frame, width=12, background='darkblue', foreground='white', borderwidth=2)

    range_date_label = tk.Label(info_frame, text="開始日付")
    start_date_entry = DateEntry(info_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
    end_date_label = tk.Label(info_frame, text="終了日付")
    end_date_entry = DateEntry(info_frame, width=12, background='darkblue', foreground='white', borderwidth=2)

    # Đặt vị trí mặc định cho ngày đơn
    single_date_label.grid(row=0, column=4, padx=10)
    single_date_entry.grid(row=0, column=5, padx=10)

    def toggle_date_mode():
        if single_date_label.winfo_ismapped():
            single_date_label.grid_remove()
            single_date_entry.grid_remove()
            range_date_label.grid(row=0, column=4, padx=10)
            start_date_entry.grid(row=0, column=5, padx=10)
            end_date_label.grid(row=0, column=6, padx=10)
            end_date_entry.grid(row=0, column=7, padx=10)
        else:
            range_date_label.grid_remove()
            start_date_entry.grid_remove()
            end_date_label.grid_remove()
            end_date_entry.grid_remove()
            single_date_label.grid(row=0, column=4, padx=10)
            single_date_entry.grid(row=0, column=5, padx=10)

    toggle_button = tk.Button(info_frame, text="切り替え", command=toggle_date_mode,bg=button_bg_color,fg="black")
    toggle_button.grid(row=0, column=9, padx=10)
    #Tao search Box

    search_label = tk.Label(info_frame, text="カーバ記号")
    search_label.grid(row=1, column=0, padx=10)
    search_entry = tk.Entry(info_frame, width=20)
    search_entry.grid(row=1, column=1, padx=10)
    search_button = tk.Button(info_frame, text="検索",  command=lambda: show_popup("search", search_entry.get()), bg="lightblue", fg="black")
    search_button.grid(row=1, column=2, padx=10)




    # Tạo bảng dữ liệu
    columns = ("カバー記号", "部位", "生産数","出荷数","基準在庫","本日在庫")

    tree_frame = tk.Frame(root)
    tree_frame.pack(expand=True, fill="both")

    scrollbar = ttk.Scrollbar(tree_frame)
    scrollbar.pack(side="right", fill="y")

    tree = ttk.Treeview(
        tree_frame,
        columns=columns,
        show="headings",
        selectmode="browse",
        yscrollcommand=scrollbar.set,
        height=10,
    )
    tree.pack(expand=True, fill="both")
    scrollbar.config(command=tree.yview)

    # Treeview columns
    tree.column("#0", anchor="w", width=120)
    tree.column(columns[0], anchor="w", width=120)
    tree.column(columns[1], anchor="w", width=120)
    tree.column(columns[2], anchor="w", width=120)
    tree.column(columns[3], anchor="w", width=120)
    tree.column(columns[4], anchor="w", width=120)



    def update_car_types(event):
        selected_factory = factory_type_combobox.get()
        if selected_factory == "平子":
            car_types = ["CRO", "SPO", "SPT", "STY", "330","全部"]
        elif selected_factory == "明治川":
            car_types = ["908", "LM", "807","全部"]
        else:
            car_types = ["CEN", "GX", "IS", "LM","LS","NX","PRA","全部"] 
        car_type_combobox['values'] = car_types
        if car_types:
            car_type_combobox.current(0)  # Đặt giá trị mặc định cho combobox loại xe

    factory_type_combobox.bind("<<ComboboxSelected>>", update_car_types)
    update_car_types(None)  # Cập nhật loại xe ban đầu


    selected_car_type = car_type_combobox.get()
    selected_factory = factory_type_combobox.get()

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor=tk.CENTER)

    
    #Ham xu ly API
    def determine_update_params(selected_factory):
        if selected_factory == "平子":
            return 'xxx', '平子'
        elif selected_factory == "明治川":
            return 'xxx', '明治川'
        else:
            return 'xxx', '明海'

    #Ham Load Data

    def load_data(popup, label, spinner):
        start_time = time.time()
        spinner.start() 

        selected_car_type = car_type_combobox.get()
       
        selected_date = single_date_entry.get_date()      
        selected_factory = factory_type_combobox.get()
        if(selected_factory=="明治川"):
            file_path = r'xxx'

        elif (selected_factory=="明海"):
            file_path = r'xxx'
        else: 
            file_path = r'xxx'


        df = pd.read_excel(file_path)
        df['車種'] = df['車種'].astype(str)
        df['車種'] = df['車種'].str.strip()
        df['カバー記号'] = df['カバー記号'].str.strip()
        df['部位'] = df['部位'].str.strip()
        df['左右'] = df['左右'].str.strip()   

        df['日付'] = df['日付'].str.strip()            
        df['日付'] = pd.to_datetime(df['日付'], format='%a, %d %b %Y %H:%M:%S %Z')   
        
        if selected_car_type == "908" and selected_factory == "明治川" :
            data = []
            count = {}
            cover_symbols = [
                "3D5", "3D6", "4D1", "4D2","5D3", "5D4","6D1", "6D2", "6D3","6D4","7D1", "7D2", "7D3", "8D1",
                 "4B1", "4B3","4B4","5B2", "5B3", "6B2", "6B3", "6B4",
                "4C1", "4C4", "5C2", "5C3","7C1", "7C2", "7C3"
            ]
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('2ndB', 'L'), ('2ndB', 'R'), ('2ndC', 'L'), ('2ndC', 'R'),('3rdB','L'),('3rdB', 'R'), ('3rdC', 'L'), ('3rdC', 'R')]
            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","2ndBL", "2ndBR", "2ndCL", "2ndCR","3rdBL", "3rdBR", "3rdCL", "3rdCR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)
            
        elif selected_car_type == "LM" and selected_factory == "明治川": 
            data = []
            count = {}
            cover_symbols = ["BR1"]

            filtered_df = df[(df['カバー記号'] == 'BR1') & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('BR1', 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['カバー記号'] == 'BR1') & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('BR1', 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == 'BR1') & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('BR1', 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['カバー記号'] == 'BR1') & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('BR1', 'FCR')] = filtered_df.shape[0] 
 
            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                if symbol in ["BR1"]:
                    part = ["FBL", "FBR", "FCL", "FCR"]           
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)

        elif selected_car_type == "807" and selected_factory == "明治川": 
            data = []
            count = {}
            cover_symbols = ["C28","3FZ"]

            filtered_df = df[(df['カバー記号'] == 'C28') & (df['部位'] == '3rdC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('C28', '3rdCL')] = filtered_df.shape[0]           
            filtered_df = df[(df['カバー記号'] == 'C28') & (df['部位'] == '3rdC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('C28', '3rdCR')] = filtered_df.shape[0]


            filtered_df = df[(df['カバー記号'] == '3FZ') & (df['部位'] == '3rdB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[('3FZ', '3rdBL')] = filtered_df.shape[0]  
            filtered_df = df[(df['カバー記号'] == '3FZ') & (df['部位'] == '3rdB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[('3FZ', '3rdBR')] = filtered_df.shape[0] 
 
            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                if symbol in ["3FZ"]:
                    part = ["3rdBL", "3rdBR"]
                else:
                    part = ["3rdCL", "3rdCR"]            
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)


        elif selected_car_type == "全部" and selected_factory == "明治川": 
            data = []
            count = {}
            cover_symbols = [
                "3D5", "3D6", "4D1", "4D2","5D3", "5D4","6D1", "6D2", "6D3","6D4","7D1", "7D2", "7D3", "8D1",
                 "4B1", "4B3","4B4","5B2", "5B3", "6B2", "6B3", "6B4",
                "4C1", "4C4", "5C2", "5C3","7C1", "7C2", "7C3",
                "C28","3FZ","BR1"
            ]
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('2ndB', 'L'), ('2ndB', 'R'), ('2ndC', 'L'), ('2ndC', 'R'),('3rdB','L'),('3rdB', 'R'), ('3rdC', 'L'), ('3rdC', 'R')]
            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","2ndBL", "2ndBR", "2ndCL", "2ndCR","3rdBL", "3rdBR", "3rdCL", "3rdCR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)

        elif selected_car_type == "330" and selected_factory == "平子":
            data = []
            count = {}
            cover_symbols = ["3XG4"]

            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row) 
        elif selected_car_type == "CRO" and selected_factory == "平子":
            data = []
            count = {}
            cover_symbols = ["8DG3", "8BD2"]

            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)    

        elif selected_car_type == "SPO" and selected_factory == "平子":
            data = []
            count = {}
            cover_symbols = ["3XG2","3XG4", "3YM3"]

            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)   
        elif selected_car_type == "SPT" and selected_factory == "平子":
            data = []
            count = {}
            cover_symbols = ["3YM3","3YQ3"]

            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)   
        elif selected_car_type == "STY" and selected_factory == "平子":
            data = []
            count = {}
            cover_symbols = ["3XG2","3XG4"]

            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row) 

        elif selected_car_type == "全部" and selected_factory == "平子":

            data = []
            count = {}
            cover_symbols = ["8DG3", "8BD2", "3XG2","3XG4","3YM3","3YQ3"]         
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)

        elif selected_car_type == "CEN" and selected_factory == "明海":
            data = []
            count = {}
            cover_symbols = ["5S0", "5S2", "5S4","6S2","6S3"]
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)



        elif selected_car_type == "GX" and selected_factory == "明海":
            data = []
            count = {}
            cover_symbols = ["3AV1", "3AV3", "3AV8","3AX1", "3BV5", "3BV9","3BX5", "3BX9", "3KW5", "3KW9",
             "3NW1", "3NW3", "3NW8", "3PZ1", "3PZ3", "3PZ8","AB1", "AB3","AD1","AD3","AD8","AN1","AN3","AN8", "AQ1","AQ3","AQ8",
             "AR1","AR3","AR8","CB1","CB3","CB8","CD1","CD3","CD8","CN1","CN3","CN8","CP1","CP3","CP8","CQ1","CQ3","CQ8","DA9","DB5","DB9",
             "DC5","DC9","DP5","DP9","DQ5","DQ9","DR5","DR9","YG1","YG3","YG8","YP1","YP3","YP8","YR1","YR3","YR8", "YW1","YW3","YW8"]

            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)

        elif selected_car_type == "IS" and selected_factory == "明海":
            data = []
            count = {}
            cover_symbols = ["WLL0","WLL2","WLL9",
             "WSL0","WSL2","WSL9","WSM0","WSM2","WSM9"] 
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)

        elif selected_car_type == "LM" and selected_factory == "明海":
            data = []
            count = {}
            cover_symbols = ["B20","B22","D10","D12","H10","H12","H20","H22","H40","H42","M30","M32","M40","M42","N20","N22",
             "R10","R12","R30","R32","V10","V12","V20","V22","V30","V32","V40","V42"] 
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)

        elif selected_car_type == "LS" and selected_factory == "明海":
            data = []
            count = {}
            cover_symbols = ["CA152","CA158","CA172","CA177","CB148","CB14H","CB157","CB15H","EB16H",
             "EB25H","KB382","KB387","KB388","KB38H","SK172","SK179"] 
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)

        elif selected_car_type == "NX" and selected_factory == "明海":
            data = []
            count = {}
            cover_symbols = ["RL8A2","RL8A9","RP8A2","RP8A9","RQ4C9","TL8A2","TL8A4","TS4B2","TS4G2","TS4G4",
             "TS4G9","TS8B2","TS8B4","TS8B9","TS8G2","TS8G4","TS8G9","TT8A2","TT8A4","TT8A9","TV8A2","TV8A4","TV8A9","TV8F2","TV8F4","TV8F9","TW4C9",
             "TW4G2","TW4G4","TW4G9","TW8C9","TW8G2","TW8G9","TX8B2","TX8B4","TX8B9","TY8B2","TY8B4","TY8B9","WR8F2","WR8F4","WR8F9","WS1B1","WS1B2",
             "WS1B4","WS1B9","WS4G1","WS5B4","WS6B2","WS8B4","WZ7A1","WZ7A2","WZ7A4","WZ7A9","WZ8A1","WZ8A2","WZ8A4","WZ8A9","XN1B2","XN4B2",
             "XN4B4","XN4B7","XN4F2","XN4F6","XN5B2","XN5B4","XN5B6","XN7A2","XN7A6","XN7A7","XN8B2","XN8B4","XN8B6","XN8B7","XN8F2","XN8F4",
             "XN8F6","XN8F7","YN4G2"] 
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]
            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)
        elif selected_car_type == "PRA" and selected_factory == "明海":
            data = []
            count = {}
            cover_symbols=["6KB2","6KW2","6MB5","6MN5","6MX2","6MX3","7AB2","7AB3","7AW2","7AW5","7BB2","7BB3","7BB5","7BX2","7BX3",
             "7KB2","7KW2","7LB2","7LW2","7LW5","7MB2","7MB5","7MN2","7PB2","7PB5","7ZB2","7ZB3","FB2","FE2","FE5","JA2","JA5","JB2","JB5","JC2",
             "JC3","JC5","JE2","JE5","JP2","JP3","XG3","XG5","XH2","XH5","XP2","XP3","XP5","XW2","XX2","XX5"] 
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)


        elif selected_car_type == "全部" and selected_factory == "明海":
            data = []
            count = {} 
            cover_symbols = ["5S0", "5S2", "5S4","6S2","6S3","3AV1", "3AV3", "3AV8","3AX1", "3BV5", "3BV9","3BX5", "3BX9", "3KW5", "3KW9",
             "3NW1", "3NW3", "3NW8", "3PZ1", "3PZ3", "3PZ8","AB1", "AB3","AD1","AD3","AD8","AN1","AN3","AN8", "AQ1","AQ3","AQ8",
             "AR1","AR3","AR8","CB1","CB3","CB8","CD1","CD3","CD8","CN1","CN3","CN8","CP1","CP3","CP8","CQ1","CQ3","CQ8","DA9","DB5","DB9",
             "DC5","DC9","DP5","DP9","DQ5","DQ9","DR5","DR9","YG1","YG3","YG8","YP1","YP3","YP8","YR1","YR3","YR8", "YW1","YW3","YW8","WLL0","WLL2","WLL9",
             "WSL0","WSL2","WSL9","WSM0","WSM2","WSM9","B20","B22","D10","D12","H10","H12","H20","H22","H40","H42","M30","M32","M40","M42","N20","N22",
             "R10","R12","R30","R32","V10","V12","V20","V22","V30","V32","V40","V42","CA152","CA158","CA172","CA177","CB148","CB14H","CB157","CB15H","EB16H",
             "EB25H","KB382","KB387","KB388","KB38H","SK172","SK179","RL8A2","RL8A9","RP8A2","RP8A9","RQ4C9","TL8A2","TL8A4","TS4B2","TS4G2","TS4G4",
             "TS4G9","TS8B2","TS8B4","TS8B9","TS8G2","TS8G4","TS8G9","TT8A2","TT8A4","TT8A9","TV8A2","TV8A4","TV8A9","TV8F2","TV8F4","TV8F9","TW4C9",
             "TW4G2","TW4G4","TW4G9","TW8C9","TW8G2","TW8G9","TX8B2","TX8B4","TX8B9","TY8B2","TY8B4","TY8B9","WR8F2","WR8F4","WR8F9","WS1B1","WS1B2",
             "WS1B4","WS1B9","WS4G1","WS5B4","WS6B2","WS8B4","WZ7A1","WZ7A2","WZ7A4","WZ7A9","WZ8A1","WZ8A2","WZ8A4","WZ8A9","XN1B2","XN4B2",
             "XN4B4","XN4B7","XN4F2","XN4F6","XN5B2","XN5B4","XN5B6","XN7A2","XN7A6","XN7A7","XN8B2","XN8B4","XN8B6","XN8B7","XN8F2","XN8F4",
             "XN8F6","XN8F7","YN4G2","6KB2","6KW2","6MB5","6MN5","6MX2","6MX3","7AB2","7AB3","7AW2","7AW5","7BB2","7BB3","7BB5","7BX2","7BX3",
             "7KB2","7KW2","7LB2","7LW2","7LW5","7MB2","7MB5","7MN2","7PB2","7PB5","7ZB2","7ZB3","FB2","FE2","FE5","JA2","JA5","JB2","JB5","JC2",
             "JC3","JC5","JE2","JE5","JP2","JP3","XG3","XG5","XH2","XH5","XP2","XP3","XP5","XW2","XX2","XX5"]      
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RC','C'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'C'),('RR2C', 'R')]

            for symbol in cover_symbols:
                for part in parts:
                    filtered_df = df[(df['カバー記号'] == symbol) & 
                                     (df['部位'] == part[0]) & 
                                     (df['左右'] == part[1]) & 
                                     (df['日付'].dt.date == selected_date)]
                    count[(symbol, part[0] + part[1])] = filtered_df.shape[0]


            total_items = sum(count.values())
            count_label.config(text=f"合計: {total_items}")                         
            for item in tree.get_children():
                tree.delete(item)
            for symbol in cover_symbols:
                part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL", "RCC","RCR","RR2BL","RR2BR","RR2CL","RR2CC","RR2CR"]
                symbol_printed = False  
                for p in part:
                    if(count.get((symbol,p),0) != 0):
                        if (symbol_printed ==False):
                            data.append((symbol, p, count.get((symbol, p), 0)))
                            symbol_printed = True  
                        else:
                            data.append(("", p, count.get((symbol, p), 0)))
            for row in data:
                tree.insert("", tk.END, values=row)
    
        end_time = time.time()
        label.config(text="完了!!!")
        spinner.stop() 
        popup.after(500, popup.destroy)  # Đóng cửa sổ popup sau 0.5 giây
        print(f"Thời gian thực thi: {end_time - start_time:.4f} giây")           
    def show_popup(task_type, keyword=None):
        # Tạo một cửa sổ con (Popup)
        popup = tk.Toplevel()
        popup.title("Loading Data")
        popup_width = 300
        popup_height = 100

        # Tính toán vị trí giữa màn hình
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x_coordinate = int((screen_width / 2) - (popup_width / 2))
        y_coordinate = int((screen_height / 2) - (popup_height / 2))

        # Định vị cửa sổ Popup ở giữa màn hình
        popup.geometry(f"{popup_width}x{popup_height}+{x_coordinate}+{y_coordinate}")

        # Thêm nhãn hiển thị trạng thái vào Popup
        label = tk.Label(popup, text="処理中...!!!", font=("Helvetica", 12))
        label.pack(pady=10)

        # Thêm thanh tiến trình vào Popupb
        spinner = ttk.Progressbar(popup, mode="indeterminate", length=100)
        spinner.pack(pady=10)

        # Chạy luồng riêng để thực hiện xử lý dữ liệu trong khi hiển thị Popup
        if task_type == "load":
            threading.Thread(target=load_data, args=(popup, label, spinner)).start()
        elif task_type == "search":
            threading.Thread(target=search_data, args=(keyword,popup, label, spinner)).start()

    #Ham lam moi Data


    def update_data(api_url,factory_name,popup, label, spinner):
        selected_factory = factory_type_combobox.get()
        url = api_url
        spinner.start() 

        try:
            # Gửi yêu cầu API và nhận dữ liệu JSON
            response = requests.get(url)
            data = response.json()

            # Chuyển đổi dữ liệu JSON thành DataFrame của pandas
            df = pd.DataFrame(data)

            # Lưu DataFrame vào một tệp Excel
            input_file = 'Input.xlsx'

            if os.path.exists(input_file):
        # Nếu tệp đã tồn tại, mở tệp và xóa tất cả dữ liệu
                workbook = load_workbook(input_file)
                sheet = workbook.active
                sheet.delete_rows(1, sheet.max_row)
                workbook.save(input_file)
            df.to_excel(input_file, index=False)
            print("Dữ liệu đã được lưu vào tệp Excel")

        except Exception as e:
            print(f"Lỗi khi lấy dữ liệu từ API: {e}")

        # Chuỗi kết nối đến SQL Server
        connection_string = (
            'DRIVER={ODBC Driver 17 for SQL Server};'
            'SERVER=ARC42618\\MEJIGAWA;'  # Thêm một dấu gạch chéo ngược bổ sung
            'DATABASE=KENSA;'
            'Trusted_Connection=yes;'
        )

        try:
            # Tạo kết nối đến SQL Server
            conn = pyodbc.connect(connection_string)
            cursor = conn.cursor()

            cursor.execute("TRUNCATE TABLE KENSA;")
            conn.commit()
            print("Dữ liệu trong bảng KENSA đã được reset thành công")
            # Load workbook
            wb = load_workbook(input_file)
            sheet = wb.active

            # Tạo DataFrame từ dữ liệu Excel
            df = pd.DataFrame(sheet.values, columns=[cell.value for cell in sheet[1]])
            df = df.iloc[1:] 
            df['日付'] = df['Date'] 
            df['工場'] = factory_name
            # Phân tích trường QR-Data
            df['カバー記号'] = df['QR-Data'].str[:6]
            df['カバー記号'] = df['カバー記号'].str.replace('-', '')
            df['部位'] = df['QR-Data'].str[8:12]
            df['部位'] = df['部位'].str.replace('-', '')
            df['左右'] = df['QR-Data'].str[12:13]
            df['連番'] = df['QR-Data'].str[13:16]
            df['車種'] = df['QR-Data'].str[16:19]
            df['車種'] = df['車種'].str.replace('-', '')
            df['縫製日'] = df['QR-Data'].str[19:25]
            df['備考'] = df['QR-Data'].str[25:]

            # Xóa cột QR-Data

            df = df.drop(columns=['QR-Data'])

            # Lưu dữ liệu vào cơ sở dữ liệu
            for index, row in df.iterrows():
                cursor.execute("""
                    INSERT INTO KENSA (カバー記号, 部位, 左右, 車種, 日付, 縫製日, 連番, 備考, 工場)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    row['カバー記号'], row['部位'], row['左右'], row['車種'], row['日付'],
                    row['縫製日'], row['連番'], row['備考'], row['工場']
                )

            # Commit các thay đổi vào cơ sở dữ liệu
            conn.commit()
            print("Dữ liệu đã được lưu vào cơ sở dữ liệu thành công")
            if(selected_factory=="明治川"):
                output_file = 'output_Meijigawa.xlsx'
            elif (selected_factory=="平子") :
                output_file = 'output_Hirako.xlsx'
            else:
                output_file = 'output_Akemi.xlsx'

            # In tất cả bảng ra file Excel
            query = "SELECT * FROM KENSA"
            df = pd.read_sql(query, conn)
            # output_file = 'output_from_sql.xlsx'
            if os.path.exists(output_file):
                try:
                    # Nếu tệp đã tồn tại, mở tệp và xóa tất cả dữ liệu
                    workbook1 = load_workbook(output_file)
                    sheet = workbook1.active
                    sheet.delete_rows(1, sheet.max_row)
                    workbook1.save(output_file)
                except Exception as e:
                    # Nếu có lỗi khi mở tệp, xóa tệp cũ
                    os.remove(output_file)
                    print(f"Tệp bị lỗi đã được xóa: {e}")

            df.to_excel(output_file, index=False)
            print(f"Dữ liệu đã được lưu vào {output_file}")

            # Xóa bảng dữ liệu (nếu cần)
            
            # Đóng kết nối
            conn.close()

        except pyodbc.Error as ex:
            sqlstate = ex.args[1]
            print(f"Lỗi kết nối đến cơ sở dữ liệu: {sqlstate}")        
        messagebox.showinfo("通知", "データが更新されました!")
        label.config(text="完了!!!")
        spinner.stop() 
        popup.after(500, popup.destroy) 


    def update_data_threaded(api_url, factory_name, root):
        # Tạo Popup và các thành phần như Label và Spinner trong hàm này
        popup = tk.Toplevel(root)
        popup.title("Loading Data")
        popup_width = 300
        popup_height = 100

        # Tính toán vị trí giữa màn hình
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x_coordinate = int((screen_width / 2) - (popup_width / 2))
        y_coordinate = int((screen_height / 2) - (popup_height / 2))

        # Định vị cửa sổ Popup ở giữa màn hình
        popup.geometry(f"{popup_width}x{popup_height}+{x_coordinate}+{y_coordinate}")

        # Thêm nhãn hiển thị trạng thái vào Popup
        label = tk.Label(popup, text="処理中...!!!", font=("Helvetica", 12))
        label.pack(pady=10)

        # Thêm thanh tiến trình vào Popup
        spinner = ttk.Progressbar(popup, mode="indeterminate", length=100)
        spinner.pack(pady=10)

        # Bắt đầu spinner
        spinner.start()

        # Tạo luồng riêng để chạy update_data
        threading.Thread(target=update_data, args=(api_url, factory_name, popup, label, spinner)).start()


    def search_data(keyword,popup,label,spinner):
        spinner.start() 
        count = {}
        data = []
        selected_date = single_date_entry.get_date()      
        selected_factory = factory_type_combobox.get()
        if(selected_factory=="明治川"):
            file_path = r'xxx'

        elif (selected_factory=="明海"):
            file_path = r'xxx'
        else: 
            file_path = r'xxx'


        df = pd.read_excel(file_path)
        df['車種'] = df['車種'].astype(str)
        df['車種'] = df['車種'].str.strip()
        df['カバー記号'] = df['カバー記号'].str.strip()
        df['部位'] = df['部位'].str.strip()
        df['左右'] = df['左右'].str.strip()   

        df['日付'] = df['日付'].str.strip()            
        df['日付'] = pd.to_datetime(df['日付'], format='%a, %d %b %Y %H:%M:%S %Z')   

        if(selected_factory == "平子" ):
            filtered_df = df[ (df['カバー記号'] == keyword) & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[(keyword, 'FBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[(keyword, 'FBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[(keyword, 'FCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[(keyword, 'FCR')] = filtered_df.shape[0] 
            filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'RB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[(keyword, 'RBL')] = filtered_df.shape[0]           
            filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'RB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[(keyword, 'RBR')] = filtered_df.shape[0]
            filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'RC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
            count[(keyword, 'RCL')] = filtered_df.shape[0]  
            filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'RC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
            count[(keyword, 'RCR')] = filtered_df.shape[0]             
            part = ["FBL", "FBR", "FCL", "FCR","RBL", "RBR", "RCL", "RCR"]       
        elif(selected_factory == "明治川"):
            if( keyword[1] == "D"):
                filtered_df = df[ (df['カバー記号'] == keyword) & (df['部位'] == '2ndB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
                count[(keyword, '2ndBL')] = filtered_df.shape[0]           
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == '2ndB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
                count[(keyword, '2ndBR')] = filtered_df.shape[0]
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == '2ndC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
                count[(keyword, '2ndCL')] = filtered_df.shape[0]  
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == '2ndC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
                count[(keyword, '2ndCR')] = filtered_df.shape[0] 
                part = ["2ndBL", "2ndBR", "2ndCL", "2ndCR"]
            elif(keyword[1] == "B"):
                filtered_df = df[ (df['カバー記号'] == keyword) & (df['部位'] == 'FB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
                count[(keyword, 'FBL')] = filtered_df.shape[0]           
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'FB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
                count[(keyword, 'FBR')] = filtered_df.shape[0]
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'FC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
                count[(keyword, 'FCL')] = filtered_df.shape[0]  
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == 'FC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
                count[(keyword, 'FCR')] = filtered_df.shape[0] 
                part = ["FBL", "FBR", "FCL", "FCR"]
            elif(keyword[1] == "C"):
                filtered_df = df[ (df['カバー記号'] == keyword) & (df['部位'] == '3rdB') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
                count[(keyword, '3rdBL')] = filtered_df.shape[0]           
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == '3rdB') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
                count[(keyword, '3rdBR')] = filtered_df.shape[0]
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == '3rdC') & (df['左右'] == 'L')& (df['日付'].dt.date == selected_date)]
                count[(keyword, '3rdCL')] = filtered_df.shape[0]  
                filtered_df = df[(df['カバー記号'] == keyword) & (df['部位'] == '3rdC') & (df['左右'] == 'R')& (df['日付'].dt.date == selected_date)]
                count[(keyword, '3rdCR')] = filtered_df.shape[0] 
                part = ["3rdBL", "3rdBR", "3rdCL", "3rdCR"]
        else:
            parts = [('FB', 'L'), ('FB', 'R'), ('FC', 'L'), ('FC', 'R'), ('RB', 'L'), ('RB', 'R'), ('RC', 'L'), ('RC', 'R'),('RR2B', 'L'), ('RR2B', 'R'), ('RR2C', 'L'), ('RR2C', 'R')]
            for part in parts:
                filtered_df = df[(df['カバー記号'] == keyword) & 
                                    (df['部位'] == part[0]) & 
                                    (df['左右'] == part[1]) & 
                                    (df['日付'].dt.date == selected_date)]
                count[(keyword, part[0] + part[1])] = filtered_df.shape[0]

            part = ["FBL", "FBR", "FCL", "FCR","RBL","RBR","RCL","RCR","RR2BL","RR2BR","RR2CL","RR2CR"]

        total_items = sum(count.values())
        count_label.config(text=f"合計: {total_items}")                         
        for item in tree.get_children():
            tree.delete(item)   
        symbol_printed = False  
        for p in part:
            if(count.get((keyword,p),0) != 0):
                if (symbol_printed ==False):
                    data.append((keyword, p, count.get((keyword, p), 0)))
                    symbol_printed = True  
                else:
                    data.append(("", p, count.get((keyword, p), 0)))
        for row in data:
            tree.insert("", tk.END, values=row)        
        label.config(text="完了!!!")
        spinner.stop() 
        popup.after(500, popup.destroy)    

    #Doan code them 2 nut 
    button_frame = tk.Frame(root)
    button_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)


    # Tạo nút 'データを読込む' và đặt vào trong button_frame
    load_data_button = tk.Button(button_frame, text="データを読込む", command=lambda: show_popup("load"),bg=button_bg_color,fg="black")
    load_data_button.pack(side=tk.LEFT, padx=5, pady=10)

    # Tạo nút 'データを更新' và đặt vào trong button_frame
    update_data_button = tk.Button(button_frame, text="データを更新",  command=lambda: update_data_threaded(*determine_update_params(factory_type_combobox.get()), root),bg=button_bg_color,fg="black")
    update_data_button.pack(side=tk.RIGHT, padx=5)


    count_label = tk.Label(root, text="合計: 0", font=("Helvetica", 12))
    count_label.pack(pady=10)
    tree.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    create_app()
